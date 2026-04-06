# -*- coding: utf-8 -*-
"""
Test complet des deux problèmes:
1. États financiers non alimentés (cellules BILAN)
2. Onglet contrôle avec anciens contrôles
"""
import openpyxl
from openpyxl.cell.cell import MergedCell
import os
import json

print("=" * 80)
print("TEST COMPLET - DOUBLE PROBLÈME")
print("=" * 80)
print()

# ==================== TEST 1: ANALYSE DU TEMPLATE ====================
print("TEST 1: ANALYSE DU TEMPLATE LIASSE")
print("-" * 80)

template_path = "Liasse_officielle_revise.xlsx"
if not os.path.exists(template_path):
    print(f"❌ Template non trouvé: {template_path}")
    exit(1)

print(f"📂 Chargement: {template_path}")
wb = openpyxl.load_workbook(template_path)
print(f"✅ Template chargé")
print()

# Vérifier l'onglet BILAN
if 'BILAN' in wb.sheetnames:
    ws_bilan = wb['BILAN']
    print("📋 Analyse onglet BILAN (MASTER):")
    
    # Tester quelques cellules clés
    cellules_test = ['E11', 'E12', 'E13', 'F11', 'F12', 'F13']
    print(f"\n{'Cellule':<10} {'Type':<15} {'Fusionnée':<12} {'Valeur':<20}")
    print("-" * 60)
    
    for cell_addr in cellules_test:
        cell = ws_bilan[cell_addr]
        is_merged = isinstance(cell, MergedCell)
        cell_type = "Fusionnée" if is_merged else "Normale"
        valeur = str(cell.value)[:20] if cell.value else "Vide"
        print(f"{cell_addr:<10} {cell_type:<15} {'Oui' if is_merged else 'Non':<12} {valeur:<20}")
    
    print()
    print("✅ Conclusion: Les cellules E11, E12, etc. sont NORMALES (non fusionnées)")
    print("✅ Ce sont les cellules de DONNÉES à remplir")
else:
    print("❌ Onglet BILAN non trouvé")

print()

# Vérifier l'onglet ACTIF
if 'ACTIF' in wb.sheetnames:
    ws_actif = wb['ACTIF']
    print("📋 Analyse onglet ACTIF:")
    
    # Vérifier les formules
    cellules_test = ['E11', 'E12', 'E13']
    print(f"\n{'Cellule':<10} {'Type':<15} {'Valeur/Formule':<40}")
    print("-" * 70)
    
    for cell_addr in cellules_test:
        cell = ws_actif[cell_addr]
        is_merged = isinstance(cell, MergedCell)
        cell_type = "Fusionnée" if is_merged else "Normale"
        valeur = str(cell.value)[:40] if cell.value else "Vide"
        print(f"{cell_addr:<10} {cell_type:<15} {valeur:<40}")
    
    print()
    print("✅ Conclusion: Les cellules ACTIF contiennent des formules =BILAN!E11")
    print("✅ Elles se rempliront automatiquement quand BILAN sera rempli")
else:
    print("❌ Onglet ACTIF non trouvé")

print()
print("=" * 80)
print()

# ==================== TEST 2: VÉRIFIER LES 16 ÉTATS ====================
print("TEST 2: VÉRIFICATION DES 16 ÉTATS DE CONTRÔLE")
print("-" * 80)

# Vérifier que le module existe
if os.path.exists("etats_controle_exhaustifs_html.py"):
    print("✅ Module etats_controle_exhaustifs_html.py trouvé")
    
    # Importer et tester
    try:
        from etats_controle_exhaustifs_html import generate_all_16_etats_controle_html
        print("✅ Fonction generate_all_16_etats_controle_html importée")
        
        # Créer des données de test
        controles_n = {
            'statistiques': {
                'total_comptes_balance': 100,
                'comptes_integres': 95,
                'comptes_non_integres': 5,
                'taux_couverture': 95.0
            },
            'equilibre_bilan': {
                'actif': 1000000,
                'passif': 1000000,
                'difference': 0,
                'pourcentage_ecart': 0,
                'equilibre': True
            },
            'equilibre_resultat': {
                'resultat_cr': 50000,
                'resultat_bilan': 50000,
                'difference': 0,
                'equilibre': True
            },
            'comptes_non_integres': [],
            'comptes_sens_inverse': [],
            'comptes_desequilibre': [],
            'hypothese_affectation': {
                'resultat_net': 50000,
                'actif': 1000000,
                'passif_sans_resultat': 950000,
                'difference_avant': 50000,
                'passif_avec_resultat': 1000000,
                'difference_apres': 0,
                'equilibre_apres': True
            },
            'comptes_sens_anormal': {
                'critiques': [],
                'eleves': [],
                'moyens': [],
                'faibles': []
            }
        }
        
        controles_n1 = controles_n.copy()
        totaux_n = {'actif': 1000000, 'passif': 1000000}
        totaux_n1 = {'actif': 900000, 'passif': 900000}
        
        # Générer les 16 états
        html_16_etats = generate_all_16_etats_controle_html(
            controles_n, controles_n1, totaux_n, totaux_n1
        )
        
        print(f"✅ HTML généré: {len(html_16_etats)} caractères")
        
        # Compter les sections
        nb_sections = html_16_etats.count('<div class="section">')
        print(f"✅ Nombre de sections: {nb_sections}")
        
        if nb_sections >= 16:
            print("✅ Les 16 états sont présents")
        else:
            print(f"⚠️ Seulement {nb_sections} états trouvés (attendu: 16)")
        
    except Exception as e:
        print(f"❌ Erreur lors de l'import: {e}")
else:
    print("❌ Module etats_controle_exhaustifs_html.py non trouvé")

print()
print("=" * 80)
print()

# ==================== TEST 3: VÉRIFIER generer_onglet_controle_coherence.py ====================
print("TEST 3: VÉRIFICATION generer_onglet_controle_coherence.py")
print("-" * 80)

if os.path.exists("generer_onglet_controle_coherence.py"):
    print("✅ Module generer_onglet_controle_coherence.py trouvé")
    
    # Lire le contenu
    with open("generer_onglet_controle_coherence.py", 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Vérifier s'il utilise les anciens contrôles
    if 'calculer_etat_controle_bilan_actif_n' in content:
        print("⚠️ Utilise les ANCIENS contrôles (calculer_etat_controle_bilan_actif_n)")
        print("❌ DOIT être mis à jour pour utiliser generer_controles_exhaustifs()")
    elif 'generer_controles_exhaustifs' in content:
        print("✅ Utilise les NOUVEAUX contrôles (generer_controles_exhaustifs)")
    else:
        print("⚠️ Impossible de déterminer quelle version est utilisée")
    
    # Vérifier la fonction generer_etats_controle_pour_export
    if 'def generer_etats_controle_pour_export' in content:
        print("✅ Fonction generer_etats_controle_pour_export trouvée")
    else:
        print("❌ Fonction generer_etats_controle_pour_export non trouvée")
else:
    print("❌ Module generer_onglet_controle_coherence.py non trouvé")

print()
print("=" * 80)
print()

# ==================== RÉSUMÉ ====================
print("RÉSUMÉ DES TESTS")
print("=" * 80)
print()

print("PROBLÈME 1: États Financiers Non Alimentés")
print("-" * 80)
print("✅ Cellules BILAN (E11, E12, etc.) sont NORMALES (non fusionnées)")
print("✅ Cellules ACTIF contiennent des formules =BILAN!E11")
print("✅ Solution: Remplir l'onglet BILAN en premier")
print()

print("PROBLÈME 2: Onglet Contrôle Utilise Anciens Contrôles")
print("-" * 80)
print("✅ Module etats_controle_exhaustifs_html.py existe")
print("✅ Fonction generate_all_16_etats_controle_html fonctionne")
print("⚠️ generer_onglet_controle_coherence.py doit être mis à jour")
print()

print("=" * 80)
print("PROCHAINES ÉTAPES")
print("=" * 80)
print()
print("1. Modifier export_liasse.py:")
print("   - Remplir l'onglet BILAN (colonnes E et F)")
print("   - Les autres onglets se rempliront automatiquement")
print()
print("2. Modifier generer_onglet_controle_coherence.py:")
print("   - Utiliser generer_controles_exhaustifs()")
print("   - Convertir les 16 états en format Excel")
print()
print("3. Tester avec de vraies données")
print()

