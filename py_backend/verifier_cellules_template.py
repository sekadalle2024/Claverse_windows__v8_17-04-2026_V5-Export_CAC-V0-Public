# -*- coding: utf-8 -*-
"""
Script pour analyser le template Liasse_officielle_revise.xlsx
et identifier les cellules exactes pour les montants
"""
import openpyxl
from openpyxl.cell.cell import MergedCell
import os

print("=" * 80)
print("ANALYSE DU TEMPLATE LIASSE OFFICIELLE")
print("=" * 80)
print()

# Charger le template
template_path = "Liasse_officielle_revise.xlsx"
if not os.path.exists(template_path):
    print(f"❌ Template non trouvé: {template_path}")
    exit(1)

print(f"📂 Chargement du template: {template_path}")
wb = openpyxl.load_workbook(template_path)
print(f"✅ Template chargé")
print()

# Analyser les onglets clés
onglets_cles = ['BILAN', 'ACTIF', 'PASSIF', 'RESULTAT']

for onglet in onglets_cles:
    if onglet not in wb.sheetnames:
        print(f"⚠️ Onglet '{onglet}' non trouvé")
        continue
    
    ws = wb[onglet]
    print(f"📋 Analyse de l'onglet: {onglet}")
    print("-" * 80)
    
    # Analyser les colonnes C, D, E, F, G (zones de montants potentielles)
    colonnes_montants = ['C', 'D', 'E', 'F', 'G']
    lignes_test = range(10, 45)  # Lignes 10 à 44 (zone de données)
    
    print(f"\n🔍 Cellules NON FUSIONNÉES (potentiellement pour les montants):")
    print(f"{'Cellule':<10} {'Type':<15} {'Valeur':<30}")
    print("-" * 80)
    
    cellules_normales = []
    for col in colonnes_montants:
        for row in lignes_test:
            cell_addr = f"{col}{row}"
            cell = ws[cell_addr]
            
            # Vérifier si c'est une cellule normale (non fusionnée)
            if not isinstance(cell, MergedCell):
                # Vérifier si la cellule est vide ou contient une formule
                valeur = cell.value
                type_cell = "Vide"
                if valeur:
                    if isinstance(valeur, str) and valeur.startswith('='):
                        type_cell = "Formule"
                    else:
                        type_cell = "Valeur"
                
                cellules_normales.append({
                    'cellule': cell_addr,
                    'type': type_cell,
                    'valeur': str(valeur)[:30] if valeur else ''
                })
    
    # Afficher les cellules normales
    for c in cellules_normales[:20]:  # Limiter à 20 pour la lisibilité
        print(f"{c['cellule']:<10} {c['type']:<15} {c['valeur']:<30}")
    
    if len(cellules_normales) > 20:
        print(f"... et {len(cellules_normales) - 20} autres cellules")
    
    print(f"\n📊 Total cellules normales: {len(cellules_normales)}")
    
    # Analyser les cellules fusionnées
    print(f"\n🔗 Cellules FUSIONNÉES (pour les titres/labels):")
    print(f"{'Range':<20} {'Cellule Principale':<20} {'Valeur':<30}")
    print("-" * 80)
    
    cellules_fusionnees = []
    for merged_range in ws.merged_cells.ranges:
        # Vérifier si la range intersecte avec nos colonnes et lignes
        for col in colonnes_montants:
            for row in lignes_test:
                cell_addr = f"{col}{row}"
                if cell_addr in merged_range:
                    top_left = merged_range.start_cell
                    valeur = top_left.value
                    cellules_fusionnees.append({
                        'range': str(merged_range),
                        'principale': top_left.coordinate,
                        'valeur': str(valeur)[:30] if valeur else ''
                    })
                    break
    
    # Afficher les cellules fusionnées (sans doublons)
    ranges_affichees = set()
    for c in cellules_fusionnees:
        if c['range'] not in ranges_affichees:
            print(f"{c['range']:<20} {c['principale']:<20} {c['valeur']:<30}")
            ranges_affichees.add(c['range'])
    
    print(f"\n📊 Total ranges fusionnées: {len(ranges_affichees)}")
    print()
    print("=" * 80)
    print()

print()
print("=" * 80)
print("RECOMMANDATIONS")
print("=" * 80)
print()
print("✅ Les cellules NON FUSIONNÉES sont les cellules de DONNÉES")
print("✅ Les cellules FUSIONNÉES sont les cellules de TITRES/LABELS")
print()
print("📝 Mettre à jour les mappings dans export_liasse.py:")
print("   - MAPPING_BILAN_ACTIF")
print("   - MAPPING_BILAN_PASSIF")
print("   - MAPPING_COMPTE_RESULTAT_CHARGES")
print("   - MAPPING_COMPTE_RESULTAT_PRODUITS")
print()
print("🔧 Utiliser les cellules normales identifiées ci-dessus")
print()

