# -*- coding: utf-8 -*-
"""
Module pour générer l'onglet "Contrôle de cohérence" dans la liasse officielle
Contient les 16 états de contrôle organisés par sections
CORRECTIONS 05 Avril 2026:
  - Bug 1: Variation N/N-1 corrigé (même liste passée deux fois → variation = 0)
  - Bug 2: Format TFT corrigé (dict converti en liste de postes)
  - Bug 3: Balance manquante gérée gracieusement
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List
import logging

logger = logging.getLogger("controle_coherence")


def format_montant_excel(montant: float) -> str:
    """Formate un montant pour Excel"""
    if abs(montant) < 0.01:
        return "-"
    return f"{montant:,.0f}".replace(',', ' ')


def ajouter_onglet_controle_coherence(wb: Workbook, etats_controle: List[Dict[str, Any]]) -> None:
    """
    Ajoute un onglet "Contrôle de cohérence" au workbook avec les 16 états de contrôle
    
    Args:
        wb: Workbook openpyxl
        etats_controle: Liste des 16 états de contrôle
    """
    logger.info("📊 Création de l'onglet 'Contrôle de cohérence'...")
    
    # Créer ou récupérer l'onglet
    if "Contrôle de cohérence" in wb.sheetnames:
        ws = wb["Contrôle de cohérence"]
        # Effacer le contenu existant
        wb.remove(ws)
    
    ws = wb.create_sheet("Contrôle de cohérence", 0)  # Insérer en première position
    
    # Styles
    style_titre_principal = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    fill_titre_principal = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    
    style_section = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    fill_section = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    style_sous_section = Font(name='Arial', size=11, bold=True, color='000000')
    fill_sous_section = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    style_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    
    style_normal = Font(name='Arial', size=10)
    style_montant = Font(name='Arial', size=10, bold=False)
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Largeurs des colonnes
    ws.column_dimensions['A'].width = 8   # REF
    ws.column_dimensions['B'].width = 50  # LIBELLÉ
    ws.column_dimensions['C'].width = 18  # EXERCICE N
    ws.column_dimensions['D'].width = 18  # EXERCICE N-1
    
    # Ligne actuelle
    row = 1
    
    # TITRE PRINCIPAL
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "CONTRÔLE DE COHÉRENCE DES ÉTATS FINANCIERS"
    cell.font = style_titre_principal
    cell.fill = fill_titre_principal
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 30
    row += 2
    
    # Sous-titre
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "16 États de Contrôle - SYSCOHADA Révisé"
    cell.font = Font(name='Arial', size=11, italic=True, color='666666')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # Organisation des états par sections
    sections = {
        'SECTION 1 : CONTRÔLES BILAN ACTIF': [1, 2, 3],
        'SECTION 2 : CONTRÔLES BILAN PASSIF': [4, 5, 6],
        'SECTION 3 : CONTRÔLES COMPTE DE RÉSULTAT': [7, 8, 9],
        'SECTION 4 : CONTRÔLES TABLEAU DES FLUX DE TRÉSORERIE': [10, 11, 12],
        'SECTION 5 : CONTRÔLES SENS DES COMPTES': [13, 14],
        'SECTION 6 : CONTRÔLES ÉQUILIBRE BILAN': [15, 16]
    }
    
    # Parcourir les sections
    for section_titre, etats_indices in sections.items():
        # Titre de section
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = section_titre
        cell.font = style_section
        cell.fill = fill_section
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        cell.border = border_thin
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Parcourir les états de cette section
        for etat_idx in etats_indices:
            if etat_idx <= len(etats_controle):
                etat = etats_controle[etat_idx - 1]
                
                # Sous-titre de l'état
                ws.merge_cells(f'A{row}:D{row}')
                cell = ws[f'A{row}']
                cell.value = etat.get('titre', f'État {etat_idx}')
                cell.font = style_sous_section
                cell.fill = fill_sous_section
                cell.alignment = Alignment(horizontal='left', vertical='center', indent=2)
                cell.border = border_thin
                ws.row_dimensions[row].height = 20
                row += 1
                
                # En-têtes des colonnes
                headers = ['REF', 'LIBELLÉ', 'EXERCICE N', 'EXERCICE N-1']
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = header
                    cell.font = style_header
                    cell.fill = fill_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border_thin
                ws.row_dimensions[row].height = 18
                row += 1
                
                # Lignes de données
                postes = etat.get('postes', [])
                for poste in postes:
                    # REF
                    cell = ws.cell(row=row, column=1)
                    cell.value = poste.get('ref', '')
                    cell.font = style_normal
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border_thin
                    
                    # LIBELLÉ
                    cell = ws.cell(row=row, column=2)
                    cell.value = poste.get('libelle', '')
                    cell.font = style_normal
                    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                    cell.border = border_thin
                    
                    # EXERCICE N
                    cell = ws.cell(row=row, column=3)
                    montant_n = poste.get('montant_n', 0) or 0
                    cell.value = format_montant_excel(montant_n)
                    cell.font = style_montant
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.border = border_thin
                    
                    # EXERCICE N-1
                    cell = ws.cell(row=row, column=4)
                    montant_n1 = poste.get('montant_n1', 0) or 0
                    cell.value = format_montant_excel(montant_n1)
                    cell.font = style_montant
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.border = border_thin
                    
                    row += 1
                
                # Ligne vide après chaque état
                row += 1
        
        # Ligne vide après chaque section
        row += 1
    
    # Pied de page
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "Note : Les montants sont exprimés en FCFA. Un tiret (-) indique un montant nul ou non applicable."
    cell.font = Font(name='Arial', size=9, italic=True, color='666666')
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30
    
    logger.info(f"✅ Onglet 'Contrôle de cohérence' créé avec {len(etats_controle)} états de contrôle")


def generer_etats_controle_pour_export(results: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Génère les 16 états de contrôle à partir des résultats des états financiers.
    
    CORRECTIONS 05 Avril 2026:
    - Bug 1: Variations N/N-1 corrigées (les 3 états de variation retournaient 0)
    - Bug 2: TFT dict converti en liste de postes pour compatibilité
    - Bug 3: Balance manquante gérée gracieusement
    
    Args:
        results: Résultats des états financiers
    
    Returns:
        Liste des 16 états de contrôle au format Excel
    """
    from etats_controle_exhaustifs import (
        calculer_etat_controle_bilan_actif_n,
        calculer_etat_controle_bilan_actif_n1,
        calculer_etat_controle_bilan_passif_n,
        calculer_etat_controle_bilan_passif_n1,
        calculer_etat_controle_compte_resultat_n,
        calculer_etat_controle_compte_resultat_n1,
        calculer_etat_controle_tft_n,
        calculer_etat_controle_tft_n1,
        calculer_etat_controle_sens_comptes_n,
        calculer_etat_controle_sens_comptes_n1,
        calculer_etat_equilibre_bilan_n,
        calculer_etat_equilibre_bilan_n1
    )
    
    logger.info("📊 Génération des 16 états de contrôle exhaustifs (version corrigée)...")
    
    try:
        # --- Extraire et normaliser les données ---
        bilan_actif = results.get('bilan_actif', [])
        bilan_passif = results.get('bilan_passif', [])
        compte_resultat = results.get('compte_resultat', [])
        balance_n_raw = results.get('balance_n', [])
        balance_n1_raw = results.get('balance_n1', [])
        tft_raw = results.get('tft', results.get('tableau_flux_tresorerie', {}))
        
        # Convertir en liste si dict
        if isinstance(bilan_actif, dict):
            bilan_actif = list(bilan_actif.values())
        if isinstance(bilan_passif, dict):
            bilan_passif = list(bilan_passif.values())
        if isinstance(compte_resultat, dict):
            compte_resultat = list(compte_resultat.values())
        
        # --- FIX BUG 3: Balances peuvent être absentes ---
        balance_n_liste = balance_n_raw if isinstance(balance_n_raw, list) else []
        balance_n1_liste = balance_n1_raw if isinstance(balance_n1_raw, list) else []
        
        # --- FIX BUG 2: Convertir le dict TFT en liste de postes ---
        # Le TFT est un dict {ZA_tresorerie_ouverture: val, ZB_flux_op: val, ...}
        # Les fonctions de contrôle TFT attendent une liste [{ref, montant_n, montant_n1}]
        tft_postes = []
        if isinstance(tft_raw, dict):
            cle_vers_ref = {
                'ZA_tresorerie_ouverture': 'ZA',
                'ZB_flux_operationnels': 'ZB',
                'ZC_flux_investissement': 'ZC',
                'ZD_flux_capitaux_propres': 'ZD',
                'ZE_flux_capitaux_etrangers': 'ZE',
                'ZF_flux_financement': 'ZF',
                'ZG_variation_tresorerie': 'ZG',
                'ZH_tresorerie_cloture': 'ZH',
                'FA_cafg': 'FA',
                'FB_variation_actif_hao': 'FB',
                'FC_variation_stocks': 'FC',
                'FD_variation_creances': 'FD',
                'FE_variation_dettes': 'FE',
            }
            for cle, ref in cle_vers_ref.items():
                if cle in tft_raw:
                    tft_postes.append({
                        'ref': ref,
                        'montant_n': float(tft_raw.get(cle, 0) or 0),
                        'montant_n1': 0
                    })
        elif isinstance(tft_raw, list):
            tft_postes = tft_raw
        
        # --- Résultat net depuis le compte de résultat ---
        resultat_net_n = 0
        resultat_net_n1 = 0
        for p in compte_resultat:
            if p.get('ref') == 'XI':
                resultat_net_n = float(p.get('montant_n', 0) or 0)
                resultat_net_n1 = float(p.get('montant_n1', 0) or 0)
                break
        if resultat_net_n == 0 and compte_resultat:
            resultat_net_n = float(compte_resultat[-1].get('montant_n', 0) or 0)
            resultat_net_n1 = float(compte_resultat[-1].get('montant_n1', 0) or 0)
        
        # --- Générer état 1 à 14 et 15, 16 via les fonctions existantes ---
        etats_controle = [
            calculer_etat_controle_bilan_actif_n(bilan_actif),                         # 1
            calculer_etat_controle_bilan_actif_n1(bilan_actif),                        # 2
            None,  # 3 - variation, calculé en dessous
            calculer_etat_controle_bilan_passif_n(bilan_passif),                       # 4
            calculer_etat_controle_bilan_passif_n1(bilan_passif),                      # 5
            None,  # 6 - variation, calculé en dessous
            calculer_etat_controle_compte_resultat_n(compte_resultat),                  # 7
            calculer_etat_controle_compte_resultat_n1(compte_resultat),                 # 8
            None,  # 9 - variation, calculé en dessous
            calculer_etat_controle_tft_n(tft_postes),                                  # 10
            calculer_etat_controle_tft_n1(tft_postes),                                 # 11
            None,  # 12 - variation TFT, calculé en dessous
            calculer_etat_controle_sens_comptes_n(balance_n_liste),                    # 13
            calculer_etat_controle_sens_comptes_n1(balance_n1_liste),                  # 14
            calculer_etat_equilibre_bilan_n(bilan_actif, bilan_passif, resultat_net_n),  # 15
            calculer_etat_equilibre_bilan_n1(bilan_actif, bilan_passif, resultat_net_n1) # 16
        ]
        
        # --- FIX BUG 1: Calculer correctement les 3 états de variation ---
        # Les fonctions originales passaient le même objet pour N et N-1
        # → variation toujours 0. On calcule directement depuis montant_n et montant_n1.
        
        total_actif_n = sum(float(p.get('montant_n', 0) or 0) for p in bilan_actif)
        total_actif_n1 = sum(float(p.get('montant_n1', 0) or 0) for p in bilan_actif)
        total_passif_n = sum(float(p.get('montant_n', 0) or 0) for p in bilan_passif)
        total_passif_n1 = sum(float(p.get('montant_n1', 0) or 0) for p in bilan_passif)
        
        # 3. Variation Bilan Actif (CORRIGÉ)
        etats_controle[2] = {
            'titre': '3. Variation Bilan Actif (N vs N-1)',
            'postes': [
                {'ref': 'CE', 'libelle': 'Total Actif Exercice N', 'montant_n': total_actif_n, 'montant_n1': 0},
                {'ref': 'CF', 'libelle': 'Total Actif Exercice N-1', 'montant_n': 0, 'montant_n1': total_actif_n1},
                {'ref': 'CG', 'libelle': 'Variation Actif (N - N-1)', 'montant_n': total_actif_n - total_actif_n1, 'montant_n1': 0},
            ]
        }
        
        # 6. Variation Bilan Passif (CORRIGÉ)
        etats_controle[5] = {
            'titre': '6. Variation Bilan Passif (N vs N-1)',
            'postes': [
                {'ref': 'PE', 'libelle': 'Total Passif Exercice N', 'montant_n': total_passif_n, 'montant_n1': 0},
                {'ref': 'PF', 'libelle': 'Total Passif Exercice N-1', 'montant_n': 0, 'montant_n1': total_passif_n1},
                {'ref': 'PG', 'libelle': 'Variation Passif (N - N-1)', 'montant_n': total_passif_n - total_passif_n1, 'montant_n1': 0},
            ]
        }
        
        # 9. Variation Compte de Résultat (CORRIGÉ)
        etats_controle[8] = {
            'titre': '9. Variation Compte de Résultat (N vs N-1)',
            'postes': [
                {'ref': 'RE', 'libelle': 'Résultat Net Exercice N', 'montant_n': resultat_net_n, 'montant_n1': 0},
                {'ref': 'RF', 'libelle': 'Résultat Net Exercice N-1', 'montant_n': 0, 'montant_n1': resultat_net_n1},
                {'ref': 'RG', 'libelle': 'Variation Résultat (N - N-1)', 'montant_n': resultat_net_n - resultat_net_n1, 'montant_n1': 0},
            ]
        }
        
        # 12. Variation TFT (CORRIGÉ)
        tft_cloture_n = next((float(p.get('montant_n', 0) or 0) for p in tft_postes if p.get('ref') == 'ZH'), 0)
        tft_variation_n = next((float(p.get('montant_n', 0) or 0) for p in tft_postes if p.get('ref') == 'ZG'), 0)
        # Pour N-1: dans le format dict, on n'a pas de N-1 direct - on utilise ZA (trésorerie ouverture de N)
        tft_ouverture_n = next((float(p.get('montant_n', 0) or 0) for p in tft_postes if p.get('ref') == 'ZA'), 0)
        
        etats_controle[11] = {
            'titre': '12. Variation Tableau des Flux de Trésorerie (N vs N-1)',
            'postes': [
                {'ref': 'TG', 'libelle': 'Trésorerie clôture N (ZH)', 'montant_n': tft_cloture_n, 'montant_n1': 0},
                {'ref': 'TH', 'libelle': 'Trésorerie ouverture N (=clôture N-1) (ZA)', 'montant_n': 0, 'montant_n1': tft_ouverture_n},
                {'ref': 'TI', 'libelle': 'Variation trésorerie nette (ZG)', 'montant_n': tft_variation_n, 'montant_n1': 0},
            ]
        }
        
        logger.info(f"   ✅ {len(etats_controle)} états de contrôle générés pour Excel (corrigés)")
        
        return etats_controle
        
    except Exception as e:
        logger.error(f"   ❌ Erreur génération états de contrôle: {e}", exc_info=True)
        etats_controle = [
            {'titre': f'{i}. État de Contrôle', 'postes': [
                {'ref': '-', 'libelle': f'Erreur: {str(e)}', 'montant_n': 0, 'montant_n1': 0}
            ]} for i in range(1, 17)
        ]
        return etats_controle
