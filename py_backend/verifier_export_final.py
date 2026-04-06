#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script de vérification finale de l'export liasse
"""
import openpyxl

print("=" * 80)
print("VÉRIFICATION FINALE EXPORT LIASSE")
print("=" * 80)

# Charger le fichier
wb = openpyxl.load_workbook('test_export_avec_correction.xlsx')

print("\n1. Onglets présents:")
for i, sheet in enumerate(wb.sheetnames[:10], 1):
    print(f"   {i}. {sheet}")

print("\n2. Vérification onglet BILAN:")
ws_bilan = wb['BILAN']
print(f"   - C10 (ACTIF N): {ws_bilan['C10'].value}")
print(f"   - D10 (ACTIF N-1): {ws_bilan['D10'].value}")
print(f"   - E10 (PASSIF N): {ws_bilan['E10'].value}")
print(f"   - F10 (PASSIF N-1): {ws_bilan['F10'].value}")

print("\n3. Vérification onglet ACTIF:")
if 'ACTIF' in wb.sheetnames:
    ws_actif = wb['ACTIF']
    c10_value = ws_actif['C10'].value
    print(f"   - C10: {c10_value}")
    if isinstance(c10_value, str) and c10_value.startswith('='):
        print(f"   ✅ C10 contient une formule")
    else:
        print(f"   ℹ️ C10 contient une valeur directe")

print("\n4. Vérification onglet Contrôle de cohérence:")
if 'Contrôle de cohérence' in wb.sheetnames:
    print("   ✅ Onglet 'Contrôle de cohérence' présent")
    ws_controle = wb['Contrôle de cohérence']
    print(f"   - Nombre de lignes: {ws_controle.max_row}")
    print(f"   - Nombre de colonnes: {ws_controle.max_column}")
else:
    print("   ❌ Onglet 'Contrôle de cohérence' absent")

print("\n5. Taille du fichier:")
import os
file_size = os.path.getsize('test_export_avec_correction.xlsx')
print(f"   - {file_size:,} bytes ({file_size/1024:.2f} KB)")

print("\n" + "=" * 80)
print("✅ VÉRIFICATION TERMINÉE")
print("=" * 80)

wb.close()
