# -*- coding: utf-8 -*-
"""
Module d'export de la liasse officielle Excel remplie avec les valeurs calculées
Version dynamique : scan des onglets pour trouver les REF SYSCOHADA.
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import io
import base64
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import Dict, Any, Optional
import logging
import datetime

logger = logging.getLogger("export_liasse")

router = APIRouter(prefix="/export-liasse", tags=["Export Liasse"])

# ==================== MODÈLES PYDANTIC ====================

class ExportLiasseRequest(BaseModel):
    """Requête pour exporter la liasse officielle"""
    results: Dict[str, Any]  # Résultats des états financiers
    nom_entreprise: Optional[str] = "ENTREPRISE"
    exercice: Optional[str] = None

class ExportLiasseResponse(BaseModel):
    success: bool
    message: str
    file_base64: Optional[str] = None
    filename: Optional[str] = None


# ==================== FONCTIONS D'EXPORT ====================

def chercher_ref_dans_feuille(ws, code_ref: str) -> list:
    """
    Cherche toutes les occurrences d'un code REF (ex: 'AD')
    dans les colonnes A, B, J, K d'une feuille Excel.
    Retourne une liste de tuples (ligne, colonne_ref).
    """
    occurrences = []
    
    # Colonnes où les REF sont susceptibles de se trouver
    cols_to_check = [1, 2, 10, 11] # A, B, J, K
    
    # Chercher sur les 100 premières lignes pour des raisons de perf
    max_row = min(ws.max_row, 150)
    for r in range(1, max_row + 1):
        for c in cols_to_check:
            cell = ws.cell(row=r, column=c)
            # Vérifier si on a la REF exacte
            if cell.value and str(cell.value).strip() == code_ref:
                occurrences.append((r, c))
                
    return occurrences

def ecrire_valeur_si_trouve(ws, ligne: int, colonne: int, montant: float):
    """
    Écrit le montant s'il y a lieu. Force l'effacement des éventuelles formules Excel.
    """
    if montant is None:
        return
        
    try:
        cell = ws.cell(row=ligne, column=colonne)
        
        # S'il y a une cellule fusionnée, il faut écrire dans la cellule en haut à gauche
        # openpyxl écrit par defaut dans la bonne cellule si l'index est le coin supérieur gauche,
        # mais sinon c'est ignoré. Dans notre cas on cible directement les colonnes H/I/J.
        
        # Forcer la valeur en écrasant toute formule
        cell.value = montant
        # Réappliquer le format nombre si nécessaire
        cell.number_format = '#,##0'
    except Exception as e:
        logger.warning(f"Erreur d'écriture à la cellule {ws.title} L{ligne}C{colonne}: {e}")

def nettoyer_montant(valeur) -> float:
    """S'assure que la valeur est un float numérique"""
    if valeur is None or valeur == '':
        return 0.0
    try:
        if isinstance(valeur, str):
            valeur = valeur.replace(' ', '').replace(',', '.')
        res = float(valeur)
        import math
        if math.isnan(res):
            return 0.0
        return res
    except (ValueError, TypeError):
        return 0.0

def injecter_donnees_dans_onglet(ws, type_onglet: str, data: list):
    """
    Injecte les données dynamiquement dans l'onglet approprié.
    Le format attendu de 'data' est une liste de dict :
    [{'ref': 'AD', 'montant_n': 100, 'montant_n1': 50, ...}]
    """
    logger.info(f"    Injection de {len(data)} postes dans l'onglet {ws.title} (Type: {type_onglet})")
    
    compteur = 0
    for poste in data:
        ref = poste.get('ref')
        if not ref:
            continue
            
        occurrences = chercher_ref_dans_feuille(ws, ref)
        if not occurrences:
            continue
            
        # Pour chaque occurrence trouvée
        for ligne, col_ref in occurrences:
            # ACTIF (Bilan ou onglet Actif dédié)
            if type_onglet == 'ACTIF':
                # Généralement REF en Col A
                # BRUT=F, AMORT=G, NET_N=H, NET_N1=I
                brut = nettoyer_montant(poste.get('brut', 0))
                amort = nettoyer_montant(poste.get('amort_deprec', 0))
                net_n = nettoyer_montant(poste.get('montant_n', poste.get('net', 0)))
                net_n1 = nettoyer_montant(poste.get('montant_n1', 0))
                
                # Sauf si on est au passif sur le Bilan complet, mais type_onglet discrimine
                
                if brut != 0:
                    ecrire_valeur_si_trouve(ws, ligne, 6, brut) # F
                if amort != 0:
                    ecrire_valeur_si_trouve(ws, ligne, 7, amort) # G
                if net_n != 0:
                    ecrire_valeur_si_trouve(ws, ligne, 8, net_n) # H
                if net_n1 != 0:
                    ecrire_valeur_si_trouve(ws, ligne, 9, net_n1) # I
                    
            # PASSIF
            elif type_onglet == 'PASSIF':
                net_n = nettoyer_montant(poste.get('montant_n', 0))
                net_n1 = nettoyer_montant(poste.get('montant_n1', 0))
                
                # Format commun: REF en A, NET_N en F ou H, NET_N1 en G ou I
                # Dans Liasse Officielle, le Bilan global a Passif souvent décalé
                # Mais si c'est un onglet PASSIF spécifique :
                # Si on est dans le Bilan consolidé, réf Passif en col J -> ecrire en M(13)/N(14)
                if col_ref >= 8: # Par ex col J (10)
                    if net_n != 0:
                        ecrire_valeur_si_trouve(ws, ligne, col_ref + 3, net_n)
                    if net_n1 != 0:
                        ecrire_valeur_si_trouve(ws, ligne, col_ref + 4, net_n1)
                else:
                    # Onglet Passif dédié
                    if net_n != 0:
                        ecrire_valeur_si_trouve(ws, ligne, 6, net_n) # F ou ajustez si besoin, essayons F=6, G=7 pour Passif seul
                    if net_n1 != 0:
                        ecrire_valeur_si_trouve(ws, ligne, 7, net_n1)
            
            # RESULTAT
            elif type_onglet == 'RESULTAT':
                net_n = nettoyer_montant(poste.get('montant_n', 0))
                net_n1 = nettoyer_montant(poste.get('montant_n1', 0))
                # Sur résultat, on pointe souvent sur N en J(10)/I(9)
                col_n = 5 if col_ref <= 2 else 6
                try: 
                    # Essayons I et J (9 et 10) qui sont standard sur CR
                    ecrire_valeur_si_trouve(ws, ligne, 9, net_n)
                    ecrire_valeur_si_trouve(ws, ligne, 10, net_n1)
                except:
                    pass
                    
            # TFT
            elif type_onglet == 'TFT':
                net_n = nettoyer_montant(poste.get('montant_n', 0))
                net_n1 = nettoyer_montant(poste.get('montant_n1', 0))
                # TFT : souvent N=I(9), N1=J(10)
                try:
                    ecrire_valeur_si_trouve(ws, ligne, 9, net_n) # I
                    ecrire_valeur_si_trouve(ws, ligne, 10, net_n1) # J
                except:
                    pass
                    
            compteur += 1
            
    logger.info(f"    -> {compteur} valeurs injectées.")


def convertir_dict_tft_vers_liste(tft_dict) -> list:
    """
    Le backend tft_v2 peut retourner un dict avec d'autres clés ('controles', etc).
    Si on nous passe l'objet global 'tft', il faut extraire la liste des postes.
    """
    if isinstance(tft_dict, list):
        return tft_dict
    elif isinstance(tft_dict, dict):
        if 'tft' in tft_dict and isinstance(tft_dict['tft'], list):
            return tft_dict['tft']
        # Ancien format où les clés étaient ZA_tresorerie...
        data_list = []
        for key, value in tft_dict.items():
            if isinstance(value, dict) and 'montant' in value:
                # Extraire la REF du début de la clé (ex: 'ZA_tresorerie_ouverture' -> 'ZA')
                ref = key[:2] if len(key) >= 2 and key[:2].isalpha() and key[:2].isupper() else key
                data_list.append({
                    'ref': ref,
                    'montant_n': value['montant'],
                    'montant_n1': value.get('montant_n1', 0)
                })
        return data_list
    return []

def remplir_liasse_officielle(results: Dict[str, Any], nom_entreprise: str, exercice: str) -> bytes:
    """
    Remplit la liasse officielle avec les valeurs calculées via scanner dynamique
    """
    logger.info("📊 Début du remplissage de la liasse officielle (Moteur Dynamique)")
    
    # Chemin du template (fichier vierge)
    # PRIORITÉ: Utiliser Liasse_officielle_revise.xlsx (84 onglets SYSCOHADA Révisé)
    template_path = "Liasse_officielle_revise.xlsx"
    if not os.path.exists(template_path):
        # Fallback vers anciens templates
        template_path = "LIASSE.xlsx"
        if not os.path.exists(template_path):
            template_path = "Liasse officielle.xlsm"
            if not os.path.exists(template_path):
                raise FileNotFoundError("Fichier template de liasse non trouvé (Liasse_officielle_revise.xlsx)")
    
    logger.info(f"📂 Template trouvé: {template_path}")
    
    # Charger le workbook
    wb = load_workbook(template_path)
    
    # Remplir Informations
    # Le comportement exact varie, mais "PAGE DE GARDE" ou "BILAN"
    
    # Bilan Actif (liste ou dict de listes format V2)
    bilan_actif_data = results.get('bilan_actif', [])
    if isinstance(bilan_actif_data, dict):
         # Cas où le routeur non-V2 nous a envoyé un dictionnaire
         bilan_actif_data = [v for k, v in bilan_actif_data.items() if isinstance(v, dict) and 'ref' in v]
         
    # Pour l'actif, on peut aussi utiliser 'actif_detaille' si disponible car il contient brut/amort
    actif_source = results.get('actif_detaille', bilan_actif_data)
    if isinstance(actif_source, dict):
         actif_source = [v for k, v in actif_source.items() if isinstance(v, dict)]

    bilan_passif_data = results.get('bilan_passif', [])
    if isinstance(bilan_passif_data, dict):
         bilan_passif_data = [v for k, v in bilan_passif_data.items() if isinstance(v, dict) and 'ref' in v]

    # Compte de Résultat
    cr_data = results.get('compte_resultat', [])
    if not cr_data:
        # Essayer de fusionner charges et produits si ancien format
        charges = results.get('charges', {})
        produits = results.get('produits', {})
        if isinstance(charges, dict):
            cr_data.extend([v for k, v in charges.items() if isinstance(v, dict)])
        if isinstance(produits, dict):
            cr_data.extend([v for k, v in produits.items() if isinstance(v, dict)])
            
    # TFT
    tft_data = convertir_dict_tft_vers_liste(results.get('tft', []))

    # Parcourir tous les onglets pour injecter de manière opportuniste
    for name in wb.sheetnames:
        ws = wb[name]
        
        # Onglets Actif
        if 'ACTIF' in name.upper() or 'BILAN' in name.upper():
            logger.info(f"→ Scan onglet {name} pour l'ACTIF")
            injecter_donnees_dans_onglet(ws, 'ACTIF', actif_source)
            
        # Onglets Passif
        if 'PASSIF' in name.upper() or 'BILAN' in name.upper():
            logger.info(f"→ Scan onglet {name} pour le PASSIF")
            injecter_donnees_dans_onglet(ws, 'PASSIF', bilan_passif_data)
            
        # Compte de Résultat
        if 'RESULTAT' in name.upper() or 'CR' in name.upper() or 'RÉSULTAT' in name.upper():
            logger.info(f"→ Scan onglet {name} pour le RÉSULTAT")
            injecter_donnees_dans_onglet(ws, 'RESULTAT', cr_data)
            
        # TFT
        if 'TFT' in name.upper() or 'TAFIRE' in name.upper():
            logger.info(f"→ Scan onglet {name} pour le TFT")
            injecter_donnees_dans_onglet(ws, 'TFT', tft_data)

    # Sauvegarder dans un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    logger.info("✅ Liasse officielle remplie avec succès (Méthode Scanner)")
    
    return output.getvalue()


# ==================== ENDPOINT API ====================

@router.post("/generer", response_model=ExportLiasseResponse)
async def generer_liasse(request: ExportLiasseRequest):
    """
    Génère la liasse officielle Excel remplie avec les valeurs calculées
    """
    try:
        logger.info("📥 Réception demande d'export liasse")
        
        # Déterminer l'exercice
        if request.exercice:
            exercice = request.exercice
        else:
            now_dt = datetime.datetime.now()
            exercice = str(now_dt.year)
        
        # Remplir la liasse
        file_content = remplir_liasse_officielle(
            results=request.results,
            nom_entreprise=request.nom_entreprise,
            exercice=exercice
        )
        
        # Encoder en base64
        file_base64 = base64.b64encode(file_content).decode('utf-8')
        
        # Nom du fichier
        nom_entreprise_propre = request.nom_entreprise.replace(' ', '_').replace('/', '_')
        filename = f"Liasse_Officielle_{nom_entreprise_propre}_{exercice}.xlsx"
        
        logger.info(f"✅ Liasse générée: {filename}")
        
        return ExportLiasseResponse(
            success=True,
            message=f"Liasse officielle générée avec succès pour {request.nom_entreprise} - Exercice {exercice}",
            file_base64=file_base64,
            filename=filename
        )
        
    except FileNotFoundError as e:
        logger.error(f"❌ Fichier template non trouvé: {e}")
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"❌ Erreur lors de la génération: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
