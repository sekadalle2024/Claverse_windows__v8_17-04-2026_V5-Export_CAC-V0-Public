# -*- coding: utf-8 -*-
"""
Script de test des corrections export liasse officielle
Test 1: Verification que les cellules ACTIF, PASSIF, RESULTAT sont remplies
Test 2: Verification que l'onglet Controle de coherence a des donnees
Test 3: Verification des 16 etats de controle
Date: 05 Avril 2026
"""

import os
import sys
import json

# Se placer dans le répertoire py_backend
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)
sys.path.insert(0, script_dir)

import openpyxl
import pandas as pd

print("=" * 70)
print("TEST CORRECTIONS EXPORT LIASSE OFFICIELLE")
print("=" * 70)
print()

erreurs = []
succes = []

# ============================================================
# ETAPE 1: Verifier les imports
# ============================================================
print("1. Vérification des imports...")
try:
    from export_liasse import remplir_liasse_officielle
    from generer_onglet_controle_coherence import (
        ajouter_onglet_controle_coherence,
        generer_etats_controle_pour_export
    )
    succes.append("✅ Imports OK")
    print("   ✅ Imports OK")
except Exception as e:
    erreurs.append(f"❌ Import error: {e}")
    print(f"   ❌ Import error: {e}")
    import traceback
    traceback.print_exc()

# ============================================================
# ETAPE 2: Charger les balances demo
# ============================================================
print()
print("2. Chargement des balances demo...")

BALANCE_FILE = "P000 -BALANCE DEMO N_N-1_N-2.xls"
balance_n = None
balance_n1 = None

if os.path.exists(BALANCE_FILE):
    try:
        balance_n = pd.read_excel(BALANCE_FILE, sheet_name=0)
        balance_n1 = pd.read_excel(BALANCE_FILE, sheet_name=1)
        print(f"   ✅ Balance N chargée: {len(balance_n)} lignes")
        print(f"   ✅ Balance N-1 chargée: {len(balance_n1)} lignes")
        succes.append(f"✅ Balances chargées depuis {BALANCE_FILE}")
    except Exception as e:
        print(f"   ⚠️ Erreur chargement {BALANCE_FILE}: {e}")
        erreurs.append(f"⚠️ Erreur chargement balance: {e}")
else:
    print(f"   ⚠️ Fichier {BALANCE_FILE} non trouvé - utilisation données fictives")

# ============================================================
# ETAPE 3: Créer des données de test
# ============================================================
print()
print("3. Création des données de test...")

# Données test représentatives format etats_financiers_v2
results_test = {
    'bilan_actif': [
        {'ref': 'AD',  'libelle': 'Charges immobilisées',     'montant_n': 1500000,  'montant_n1': 1200000},
        {'ref': 'AE',  'libelle': 'Frais de recherche',       'montant_n': 800000,   'montant_n1': 700000},
        {'ref': 'AF',  'libelle': 'Brevets, licences',        'montant_n': 2000000,  'montant_n1': 1800000},
        {'ref': 'AI',  'libelle': 'Terrains',                 'montant_n': 5000000,  'montant_n1': 5000000},
        {'ref': 'AJ',  'libelle': 'Batiments',                'montant_n': 12000000, 'montant_n1': 13000000},
        {'ref': 'AL',  'libelle': 'Materiel',                 'montant_n': 4500000,  'montant_n1': 5000000},
        {'ref': 'AZ',  'libelle': 'Total Actif Immobilisc',   'montant_n': 25800000, 'montant_n1': 26700000},
        {'ref': 'BB',  'libelle': 'Stocks et encours',        'montant_n': 3200000,  'montant_n1': 2800000},
        {'ref': 'BI',  'libelle': 'Clients',                  'montant_n': 4800000,  'montant_n1': 5200000},
        {'ref': 'BQ',  'libelle': 'Total Actif Circulant',    'montant_n': 8000000,  'montant_n1': 8000000},
        {'ref': 'BV',  'libelle': 'Banques, cheques, caisse', 'montant_n': 2200000,  'montant_n1': 1800000},
        {'ref': 'BZ',  'libelle': 'Total Tresorerie Actif',   'montant_n': 2200000,  'montant_n1': 1800000},
        {'ref': 'DZ',  'libelle': 'TOTAL ACTIF',              'montant_n': 36000000, 'montant_n1': 36500000},
    ],
    'bilan_passif': [
        {'ref': 'DA',  'libelle': 'Capital',                  'montant_n': 10000000, 'montant_n1': 10000000},
        {'ref': 'DF',  'libelle': 'Reserves libres',          'montant_n': 3500000,  'montant_n1': 2500000},
        {'ref': 'DH',  'libelle': 'Resultat net',             'montant_n': 1500000,  'montant_n1': 1200000},
        {'ref': 'DZ',  'libelle': 'Total Capitaux Propres',   'montant_n': 15000000, 'montant_n1': 13700000},
        {'ref': 'RA',  'libelle': 'Emprunts',                 'montant_n': 8000000,  'montant_n1': 9500000},
        {'ref': 'RZ',  'libelle': 'Total Dettes Financieres', 'montant_n': 8000000,  'montant_n1': 9500000},
        {'ref': 'TC',  'libelle': 'Fournisseurs exploitation', 'montant_n': 7500000, 'montant_n1': 7800000},
        {'ref': 'TD',  'libelle': 'Dettes fiscales',          'montant_n': 3200000,  'montant_n1': 3200000},
        {'ref': 'TZ',  'libelle': 'Total Passif Circulant',   'montant_n': 13000000, 'montant_n1': 13300000},
    ],
    'compte_resultat': [
        {'ref': 'RA',  'libelle': 'Ventes de marchandises',   'montant_n': 25000000, 'montant_n1': 22000000},
        {'ref': 'RB',  'libelle': 'Ventes produits fabriques','montant_n': 8000000,  'montant_n1': 7500000},
        {'ref': 'RZ',  'libelle': 'Total Produits Exploitation','montant_n': 33000000,'montant_n1': 29500000},
        {'ref': 'TA',  'libelle': 'Achats de marchandises',   'montant_n': 15000000, 'montant_n1': 13500000},
        {'ref': 'TK',  'libelle': 'Charges de personnel',     'montant_n': 9000000,  'montant_n1': 8500000},
        {'ref': 'TL',  'libelle': 'Dotations amortissements', 'montant_n': 2500000,  'montant_n1': 2300000},
        {'ref': 'TZ',  'libelle': 'Total Charges Exploitation','montant_n': 28500000, 'montant_n1': 26000000},
        {'ref': 'XI',  'libelle': 'RESULTAT NET',             'montant_n': 1500000,  'montant_n1': 1200000},
    ],
    'tft': {
        'ZA_tresorerie_ouverture': 1800000,
        'FA_cafg': 4000000,
        'ZB_flux_operationnels': 3200000,
        'FF_decaissement_incorp': -500000,
        'FG_decaissement_corp': -1000000,
        'ZC_flux_investissement': -1500000,
        'FK_augmentation_capital': 0,
        'ZD_flux_capitaux_propres': 0,
        'FO_nouveaux_emprunts': 0,
        'FQ_remboursements': -1300000,
        'ZE_flux_capitaux_etrangers': -1300000,
        'ZF_flux_financement': -1300000,
        'ZG_variation_tresorerie': 400000,
        'ZH_tresorerie_cloture': 2200000,
    },
}

print(f"   ✅ Données test créées")
print(f"      - Bilan Actif: {len(results_test['bilan_actif'])} postes")
print(f"      - Bilan Passif: {len(results_test['bilan_passif'])} postes")
print(f"      - Compte Résultat: {len(results_test['compte_resultat'])} postes")

# ============================================================
# ETAPE 4: Tester la génération des états de contrôle
# ============================================================
print()
print("4. Test génération états de contrôle...")
try:
    from generer_onglet_controle_coherence import generer_etats_controle_pour_export
    etats = generer_etats_controle_pour_export(results_test)
    
    print(f"   ✅ {len(etats)} états générés")
    
    # Vérifier les variations (états 3, 6, 9, 12)
    for idx, nom in [(2, 'Variation Actif'), (5, 'Variation Passif'), (8, 'Variation CR')]:
        etat = etats[idx]
        # Chercher un poste de variation
        variation_postes = [p for p in etat.get('postes', []) if 'Variation' in p.get('libelle', '')]
        if variation_postes:
            v = variation_postes[0]['montant_n']
            statut = "✅" if v != 0 else "⚠️ (0 - peut être normal si N==N-1)"
            print(f"   {statut} État {idx+1} ({nom}): variation = {v:,.0f}")
        else:
            print(f"   ℹ️ État {idx+1} ({nom}): pas de poste 'Variation' trouvé")
    
    # Vérifier les totaux
    etat15 = etats[14]  # Équilibre bilan N
    for p in etat15.get('postes', []):
        if p.get('ref') == 'EA':
            print(f"   ✅ État 15 - Total Actif N: {p['montant_n']:,.0f}")
    
    succes.append(f"✅ {len(etats)} états de contrôle générés")
    
except Exception as e:
    erreurs.append(f"❌ Erreur états contrôle: {e}")
    print(f"   ❌ Erreur: {e}")
    import traceback
    traceback.print_exc()

# ============================================================
# ETAPE 5: Tester l'export liasse officielle
# ============================================================
print()
print("5. Test export liasse officielle...")

template_path = "Liasse_officielle_revise.xlsx"
if not os.path.exists(template_path):
    print(f"   ⚠️ Template {template_path} non trouvé - test simplifié")
    erreurs.append(f"⚠️ Template liasse non trouvé: {template_path}")
else:
    try:
        file_bytes = remplir_liasse_officielle(results_test, "ENTREPRISE TEST", "2024")
        
        # Sauvegarder
        output = "test_corrections_export_v2.xlsx"
        with open(output, 'wb') as f:
            f.write(file_bytes)
        print(f"   ✅ Fichier Excel généré: {output} ({len(file_bytes):,} bytes)")
        
        # Vérifier le contenu
        wb_check = openpyxl.load_workbook(output, data_only=True)
        print(f"   📋 Onglets: {wb_check.sheetnames[:8]}")
        
        # Vérifier onglet ACTIF
        cellules_actif = 0
        if 'ACTIF' in wb_check.sheetnames:
            ws = wb_check['ACTIF']
            for r in range(10, 45):
                # N est en H
                v = ws[f'H{r}'].value
                if v is not None and v != 0 and v != '-':
                    cellules_actif += 1
            print(f"   {'✅' if cellules_actif > 0 else '❌'} Onglet ACTIF: {cellules_actif} montants pertinents ecrits en col H")
            if cellules_actif == 0:
                erreurs.append(f"❌ Onglet ACTIF vide")
            else:
                succes.append(f"✅ Onglet ACTIF: {cellules_actif} montants ecrits")
        
        # Vérifier onglet PASSIF
        cellules_passif = 0
        if 'PASSIF' in wb_check.sheetnames:
            ws = wb_check['PASSIF']
            for r in range(10, 45):
                v = ws[f'H{r}'].value  # N est en H
                if v is not None and v != 0 and v != '-':
                    cellules_passif += 1
            print(f"   {'✅' if cellules_passif > 0 else '❌'} Onglet PASSIF: {cellules_passif} montants pertinents ecrits en col H")
            if cellules_passif == 0:
                erreurs.append(f"❌ Onglet PASSIF vide")
            else:
                succes.append(f"✅ Onglet PASSIF: {cellules_passif} montants ecrits")
                
        # Vérifier onglet RESULTAT
        cellules_cr = 0
        if 'RESULTAT' in wb_check.sheetnames:
            ws = wb_check['RESULTAT']
            for r in range(10, 45):
                v = ws[f'I{r}'].value  # N est en I
                if v is not None and v != 0 and v != '-':
                    cellules_cr += 1
            print(f"   {'✅' if cellules_cr > 0 else '❌'} Onglet RESULTAT: {cellules_cr} montants pertinents ecrits en col I")
            if cellules_cr == 0:
                erreurs.append(f"❌ Onglet RESULTAT vide")
            else:
                succes.append(f"✅ Onglet RESULTAT: {cellules_cr} montants ecrits")
                
        # Vérifier onglet BILAN
        cellules_bilan = 0
        if 'BILAN' in wb_check.sheetnames:
            ws = wb_check['BILAN']
            for r in range(10, 45):
                if ws[f'H{r}'].value or ws[f'M{r}'].value:
                    if ws[f'H{r}'].value != 0 and ws[f'M{r}'].value != 0:
                        cellules_bilan += 1
            print(f"   {'✅' if cellules_bilan > 0 else '❌'} Onglet BILAN: {cellules_bilan} montants pertinents ecrits")
            if cellules_bilan == 0:
                erreurs.append(f"❌ Onglet BILAN vide")
            else:
                succes.append(f"✅ Onglet BILAN: {cellules_bilan} montants ecrits")
                
        # Vérifier onglet TFT
        cellules_tft = 0
        if 'TFT' in wb_check.sheetnames:
            ws = wb_check['TFT']
            for r in range(10, 45):
                if ws[f'H{r}'].value and ws[f'H{r}'].value != 0:
                    cellules_tft += 1
            print(f"   {'✅' if cellules_tft > 0 else '❌'} Onglet TFT: {cellules_tft} montants pertinents ecrits")
            if cellules_tft == 0:
                erreurs.append(f"❌ Onglet TFT vide")
            else:
                succes.append(f"✅ Onglet TFT: {cellules_tft} montants ecrits")
        # Vérifier onglet Contrôle de cohérence
        if "Contrôle de cohérence" in wb_check.sheetnames:
            ws_ctrl = wb_check["Contrôle de cohérence"]
            print(f"   ✅ Onglet 'Contrôle de cohérence' présent ({ws_ctrl.max_row} lignes)")
            succes.append("✅ Onglet Contrôle de cohérence présent")
        else:
            print(f"   ⚠️ Onglet 'Contrôle de cohérence' absent")
            erreurs.append("❌ Onglet Contrôle de cohérence absent")
        
        wb_check.close()
        
    except Exception as e:
        erreurs.append(f"❌ Erreur export liasse: {e}")
        print(f"   ❌ Erreur: {e}")
        import traceback
        traceback.print_exc()

# ============================================================
# RAPPORT FINAL
# ============================================================
print()
print("=" * 70)
print("RAPPORT FINAL")
print("=" * 70)
print(f"✅ Succès: {len(succes)}")
for s in succes:
    print(f"   {s}")
print()
if erreurs:
    print(f"❌/⚠️ Problèmes: {len(erreurs)}")
    for e in erreurs:
        print(f"   {e}")
else:
    print("🎉 TOUS LES TESTS PASSÉS!")
print("=" * 70)
