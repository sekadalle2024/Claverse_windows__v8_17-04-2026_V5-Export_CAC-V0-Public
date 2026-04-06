"""
Scanner les onglets du template pour trouver les vraies positions REF → cellule
"""
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, column_index_from_string
import os
os.chdir(os.path.dirname(os.path.abspath(__file__)))

wb = openpyxl.load_workbook('Liasse_officielle_revise.xlsx')

def trouver_ref_mapping(ws, col_ref='A', cols_data=None):
    """
    Scanne une feuille Excel et retourne un dict {ref: [(row, col_data), ...]}.
    Cherche les codes REF SYSCOHADA (2 lettres majuscules) dans la colonne col_ref.
    Puis trouve les premières cellules NON-MergedCell dans cols_data sur la même ligne.
    """
    mapping = {}
    col_ref_idx = column_index_from_string(col_ref)
    
    for row in ws.iter_rows():
        ref_cell = ws.cell(row[0].row, col_ref_idx)
        ref_val = str(ref_cell.value or '').strip()
        
        # Vérifier si c'est un code REF SYSCOHADA (2 lettres majuscules)
        if len(ref_val) == 2 and ref_val.isalpha() and ref_val.isupper():
            row_num = ref_cell.row
            
            # Trouver les colonnes de données disponibles (non-MergedCell)
            data_cells = []
            if cols_data:
                for col_letter in cols_data:
                    col_idx = column_index_from_string(col_letter)
                    cell = ws.cell(row_num, col_idx)
                    if not isinstance(cell, MergedCell):
                        data_cells.append(f'{col_letter}{row_num}')
                    else:
                        # Trouver le top-left
                        for mr in ws.merged_cells.ranges:
                            if cell.coordinate in mr:
                                top_left = f'{get_column_letter(mr.min_col)}{mr.min_row}'
                                if top_left not in data_cells:
                                    data_cells.append(f'[MERGE:{top_left}]')
                                break
            
            mapping[ref_val] = {'row': row_num, 'cells': data_cells}
    
    return mapping

print('=== ACTIF - REF en col A, data en cols C,D,E,F ===')
ws = wb['ACTIF']
mapping = trouver_ref_mapping(ws, 'A', ['C', 'D', 'E', 'F'])
for ref, info in sorted(mapping.items()):
    print(f'  {ref}: row {info["row"]:3d} → {info["cells"]}')

print()
print('=== PASSIF - REF en col A, data en cols G,H,I ===')
ws = wb['PASSIF']
# Trouver les colonnes non-mergées dans PASSIF
# D'abord scanner la row 11 pour voir les colonnes dispo
ws_p = wb['PASSIF']
row11_info = []
for col_idx in range(1, 15):
    cell = ws_p.cell(11, col_idx)
    t = 'Cell' if not isinstance(cell, MergedCell) else 'Merged'
    row11_info.append(f'{get_column_letter(col_idx)}={t}(val={cell.value})')
print('  Row 11 types:', ', '.join(row11_info))

print()
mapping = trouver_ref_mapping(ws_p, 'A', ['G', 'H', 'I', 'J'])
for ref, info in sorted(mapping.items()):
    print(f'  {ref}: row {info["row"]:3d} → {info["cells"]}')

print()
print('=== RESULTAT - REF en col A, data en cols G,H,I,J ===')
ws = wb['RESULTAT']
# Idem, trouver colonnes dispo
ws_r = wb['RESULTAT']
row11_info = []
for col_idx in range(1, 15):
    cell = ws_r.cell(11, col_idx)
    t = 'Cell' if not isinstance(cell, MergedCell) else 'Merged'
    v = str(cell.value or '')[:10]
    row11_info.append(f'{get_column_letter(col_idx)}={t}({v})')
print('  Row 11 types:', ', '.join(row11_info))

print()
mapping = trouver_ref_mapping(ws_r, 'A', ['F', 'G', 'H', 'I', 'J'])
for ref, info in sorted(mapping.items()):
    print(f'  {ref}: row {info["row"]:3d} → {info["cells"]}')
